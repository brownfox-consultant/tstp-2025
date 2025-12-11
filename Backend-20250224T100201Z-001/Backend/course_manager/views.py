import logging
import uuid

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import status, viewsets
from rest_framework.decorators import action, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.decorators import api_view
from django.utils.timezone import now, timedelta
from sTest.aws_client import AwsStorageClient
from sTest.permissions import IsAdminOrContentDeveloperOrFaculty, IsAdminOrContentDeveloper, IsAdmin, \
    IsAdminOrContentDeveloperOrFacultyOrStudent
from sTest.utils import get_error_response_for_serializer, get_error_response, CustomPageNumberPagination
from user_manager.serializers import StudentSerializer
from system_manager.models import Suggestion
from .filters import QuestionFilter, MaterialFilter
from .models import Question, QuestionLog ,Course, Subject, CourseSubjects, Material, CourseEnrollment, Topic
from .serializers import CreateQuestionSerializer, CourseWithSubjectsSerializer, QuestionListSerializer, \
    CourseSerializer, CreateCourseSerializer, MaterialSerializer, SubjectSerializer, \
    MaterialListSerializer, MaterialDetailsSerializer, TopicSerializer
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import ValidationError
from django.db.models import Max 
import csv
from django.http import HttpResponse
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser
from django.db.models import Count
from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework.decorators import action
import datetime
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
import datetime
from django_filters.rest_framework import DjangoFilterBackend
from .filters import CourseFilter


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.get_all()
    serializer_class = CourseSerializer
    logger = logging.getLogger('Courses')
    filter_backends = [DjangoFilterBackend]
    filterset_class = CourseFilter

    @action(detail=True, methods=['get'], permission_classes=[AllowAny], url_path='subjects')
    def get_subjects_by_course(self, request, pk=None):
        """
        Get all subjects for a given course ID
        GET /api/course/{course_id}/subjects/
        """
        try:
            course = Course.get_course_by_id(course_id=pk)
            course_subjects = CourseSubjects.objects.filter(course=course)
            serializer = SubjectSerializer([cs.subject for cs in course_subjects], many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Course.DoesNotExist:
            return Response({'error': 'Course not found'}, status=status.HTTP_404_NOT_FOUND)


    @action(detail=False, methods=['get'], url_path='constants')
    def get_question_constants(self, request):
        # Get choices from Question model
        difficulty_list = [
            choice[0] for choice in Question.DIFFICULTY_CHOICES
            if choice[0] in ['EASY', 'MODERATE', 'HARD']
        ]
        question_type_list = [choice[0] for choice in Question.QUESTION_TYPE]
        test_type_list = [choice[0] for choice in Question.TEST_TYPE_CHOICES]
        question_subtype_list = [
            {"value": choice[0], "label": choice[1]}
            for choice in Question.QUESTION_SUBTYPE
        ]


        # If you also want topics:
        topics = Topic.objects.all()
        serialized_topics = TopicSerializer(topics, many=True).data

        return Response({
            "topics": serialized_topics,
            "difficultyList": difficulty_list,
            "questionTypeList": question_type_list,
            "testTypeList": test_type_list,
            "questionSubtypeList": question_subtype_list,  # ✅ added
        })

        
    @permission_classes([IsAdmin])
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = CreateCourseSerializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
            data = serializer.validated_data

            # Create or get the Course
            course, _ = Course.objects.get_or_create(name=data['name'])

            for subject_data in data['subjects']:
                # Create or get the Subject
                subject, _ = Subject.objects.get_or_create(name=subject_data['name'])

                # Append an ID to each section
                for index, section in enumerate(subject_data['sections']):
                    section['id'] = index + 1  # ID based on index

                # Prepare metadata for CourseSubjects
                metadata = {'sections': subject_data['sections']}

                # Create CourseSubjects entry
                CourseSubjects.objects.create(
                    course=course,
                    subject=subject,
                    metadata=metadata,
                    correct_answer_marks=subject_data['correct_answer_marks'],
                    incorrect_answer_marks=subject_data['incorrect_answer_marks'],
                    order=subject_data['order']
                )

            return Response({"detail": "Course and subjects created successfully"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return get_error_response_for_serializer(logger=self.logger, serializer=serializer, data=request.data)

    @permission_classes([IsAdminOrContentDeveloperOrFaculty])
    def list(self, request, *args, **kwargs):
        courses = Course.get_all()
        serializer = CourseWithSubjectsSerializer(courses, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @permission_classes([IsAdmin])
    def retrieve(self, request, pk=None, *args, **kwargs):
        course = Course.get_course_by_id(course_id=pk)
        serializer = CourseWithSubjectsSerializer(course)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @permission_classes([IsAdmin])
    @transaction.atomic
    def update(self, request, pk=None, *args, **kwargs):
        course = Course.get_course_by_id(course_id=pk)
        serializer = CreateCourseSerializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
        except Exception as e:
            return get_error_response_for_serializer(logger=self.logger, serializer=serializer, data=request.data)

        data = serializer.validated_data

        # Update the course name if it has changed
        if course.name != data['name']:
            course.name = data['name']
            course.updated_at = timezone.now()
            course.save()

        existing_subjects = CourseSubjects.objects.filter(course=course)
        for subject_data in data['subjects']:
            self.logger.info(f'Processing subject {subject_data} for course {course.id}')
            subject, _ = Subject.objects.get_or_create(name=subject_data['name'])

            # Append an ID to each section
            for index, section in enumerate(subject_data['sections']):
                section['id'] = index + 1

            # Update metadata for CourseSubjects
            metadata = {'sections': subject_data['sections']}

            course_subject = None
            if 'course_subject_id' in subject_data:
                course_subject = CourseSubjects.objects.get(id=subject_data['course_subject_id'])
            else:
                course_subject = CourseSubjects.objects.filter(course=course.id, subject=subject.id).first()

            # Check if we are updating an existing subject in this course
            if course_subject is not None:
                course_subject.course = course
                course_subject.subject = subject
                course_subject.metadata = metadata
                course_subject.correct_answer_marks = subject_data['correct_answer_marks']
                course_subject.incorrect_answer_marks = subject_data['incorrect_answer_marks']
                course_subject.order = subject_data['order']
                course_subject.save()
            else:
                CourseSubjects.objects.create(
                    course=course,
                    subject=subject,
                    metadata=metadata,
                    correct_answer_marks=subject_data['correct_answer_marks'],
                    incorrect_answer_marks=subject_data['incorrect_answer_marks'],
                    order=subject_data['order']
                )
            existing_subjects = existing_subjects.exclude(Q(subject=subject))

        # Delete any remaining old subjects
        existing_subjects.delete()
        return Response({"detail": "Course and subjects updated successfully"}, status=status.HTTP_200_OK)

    @permission_classes([IsAdmin])
    def destroy(self, request, pk=None, *args, **kwargs):
        instance = Course.get_course_by_id(course_id=pk)
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_destroy(self, instance):
        # instance.is_active = False
        # instance.updated_at = timezone.now()
        # instance.save()
        instance.delete()

    @action(detail=True, methods=['patch'], permission_classes=[IsAdmin], url_path='deactivate')
    def deactivate_course(self, request, pk=None):
        course = Course.get_course_by_id(course_id=pk)
        # course.is_active = False
        # course.updated_at = timezone.now()
        # course.save()
        course.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['GET'], permission_classes=[AllowAny], url_path='list')
    def list_courses(self, request):
        courses = Course.get_all()
        serializer = CourseSerializer(courses, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['GET'], permission_classes=[IsAdminOrContentDeveloperOrFaculty],
            url_path='(?P<course_subject_id>\d+)/questions')
    def list_questions_by_subject(self, request, course_subject_id=None):
        if not course_subject_id:
            self.logger.exception('Error processing the request because no course subject id was provided')
            return get_error_response('Subject is mandatory')

        questions = Question.get_questions_for_subject(course_subject_id=course_subject_id)
        topics = Topic.objects.filter(course_subject_id=course_subject_id)

        # Apply dynamic filtering
        filter_backends = [DjangoFilterBackend]
        filterset = QuestionFilter(request.GET, queryset=questions)
        if not filterset.is_valid():
            return get_error_response('Invalid filter parameters')

        filtered_questions = filterset.qs

        # Apply pagination
        paginator = CustomPageNumberPagination()
        paginator.page_size = 15
        paginated_questions = paginator.paginate_queryset(filtered_questions, request)

        questions_serializer = QuestionListSerializer(paginated_questions, many=True)
        topics_serializer = TopicSerializer(topics, many=True)

        # Return the paginated response with topics
        return paginator.get_paginated_response({
            'questions': questions_serializer.data,
            'topics': topics_serializer.data
        })

    @action(detail=False, methods=['GET'], permission_classes=[IsAdminOrContentDeveloperOrFacultyOrStudent],
            url_path='(?P<course_subject_id>\d+)/topics')
    def list_topics_by_subject(self, request, course_subject_id=None):
        if not course_subject_id:
            self.logger.exception('Error processing the request because no course subject id was provided')
            return get_error_response('Subject is mandatory')

        topics = Topic.objects.filter(course_subject_id=course_subject_id)

        topics_serializer = TopicSerializer(topics, many=True)

        return Response(data=topics_serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['GET'], url_path='students')
    def list_students_for_course(self, request, pk=None, *args, **kwargs):
        course = Course.get_course_by_id(course_id=pk)
        students = course.get_enrolled_students()

        # Apply pagination
        paginator = CustomPageNumberPagination()
        paginator.page_size = 15
        paginated_students = paginator.paginate_queryset(students, request)

        # Serialize page of students
        serializer = StudentSerializer(paginated_students, many=True)

        # Return the paginated response
        return paginator.get_paginated_response(serializer.data)


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.get_all()
    serializer_class = SubjectSerializer
    permission_classes = [IsAdmin]

    def create(self, request, *args, **kwargs):
        return super(SubjectViewSet, self).create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        return super(SubjectViewSet, self).update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_destroy(self, instance):
        # instance.is_active = False
        # instance.updated_at = timezone.now()
        # instance.save()
        instance.delete()

    @action(detail=True, methods=['patch'], permission_classes=[IsAdmin], url_path='deactivate')
    def deactivate_subject(self, request, pk=None):
        try:
            subject = self.get_object()
            # subject.is_active = False
            # subject.updated_at = timezone.now()
            # subject.save()
            subject.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Subject.DoesNotExist:
            return Response({'detail': 'Subject not found.'}, status=status.HTTP_404_NOT_FOUND)


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 200

class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.get_all()
    serializer_class = CreateQuestionSerializer
    logger = logging.getLogger('Questions')
    pagination_class = StandardResultsSetPagination

    @staticmethod
    def get_client_ip(request):
        """Helper to extract client IP address safely."""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]  # take first if multiple
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

    @action(detail=False, methods=['get'], url_path='export-by-subtopics', permission_classes=[IsAdminOrContentDeveloper])
    def export_by_subtopics(self, request):
        """
        Export questions for given sub_topic_ids and course_subject_id into CSV.
        Example: ?course_subject_id=10&sub_topic_ids=1,21,3,5,6
        """
        course_subject_id = request.query_params.get("course_subject_id")
        sub_topic_ids = request.query_params.get("sub_topic_ids")

        if not course_subject_id or not sub_topic_ids:
            return Response(
                {"error": "Both course_subject_id and sub_topic_ids are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            sub_topic_ids = [int(x) for x in sub_topic_ids.split(",") if x.strip()]
        except ValueError:
            return Response({"error": "sub_topic_ids must be a comma-separated list of integers."},
                            status=status.HTTP_400_BAD_REQUEST)

        # ✅ Fetch questions
        questions = Question.objects.filter(
            course_subject_id=course_subject_id,
            sub_topic_id__in=sub_topic_ids
        ).select_related(
            "course_subject__course",
            "course_subject__subject",
            "sub_topic",
            "topic",
        )

        # ✅ CSV Response
        response = HttpResponse(content_type="text/csv")
        filename = f"questions_by_subtopics_{datetime.datetime.now().strftime('%Y%m%d')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow([
            "ID", "Course", "Subject", "Topic", "Sub Topic", "Description",
            "Question Type", "Subtype", "Options", "Difficulty", "Test Type",
            "Directions", "Explanation", "Srno", "Created At", "Updated At"
        ])

        for q in questions:
            writer.writerow([
                q.id,
                q.course_subject.course.name if q.course_subject else "",
                q.course_subject.subject.name if q.course_subject else "",
                q.topic.name if q.topic else "",
                q.sub_topic.name if q.sub_topic else "",   # ✅ sub_topic name instead of id
                q.description,
                q.question_type,
                q.question_subtype,
                q.options,
                q.difficulty,
                q.test_type,
                q.directions,
                q.explanation,
                q.srno,
                q.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                q.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            ])

        return response

    @action(detail=False, methods=['get'], url_path='duplicates')
    def find_duplicates(self, request):
        """
        Find duplicate questions (same description + options).
        Supports ?export=csv or ?export=xlsx.
        Shows 2-row break between duplicate groups.
        """
        import csv, json, datetime
        from openpyxl import Workbook
        from django.db.models import Count
        from django.http import HttpResponse
        from collections import defaultdict

        # Step 1: Find duplicates by description + options
        duplicates = (
            Question.objects.values("description", "options")
            .annotate(count=Count("id"))
            .filter(count__gt=1)
        )

        # Step 2: Get all matching questions
        duplicate_questions = (
            Question.objects.filter(
                description__in=[d["description"] for d in duplicates],
            )
            .select_related("course_subject__course", "course_subject__subject")
            .order_by("description", "srno")
        )

        # Step 3: Group by (description, options as JSON string)
        grouped = defaultdict(list)
        for q in duplicate_questions:
            try:
                options_key = json.dumps(q.options, sort_keys=True) if isinstance(q.options, (list, dict)) else str(q.options)
            except Exception:
                options_key = str(q.options)
            grouped[(q.description, options_key)].append(q)

        export_format = request.GET.get("export")

        # ===== CSV EXPORT =====
        if export_format == "csv":
            response = HttpResponse(content_type="text/csv")
            response["Content-Disposition"] = 'attachment; filename="duplicate_questions.csv"'
            writer = csv.writer(response)
            writer.writerow(["Srno", "Course", "Subject", "Question Type", "Subtype", "Difficulty", "Test Type"])

            for _, questions in grouped.items():
                if len(questions) < 2:
                    continue  # skip non-duplicates
                for q in questions:
                    writer.writerow([
                        q.srno,
                        q.course_subject.course.name if q.course_subject else "",
                        q.course_subject.subject.name if q.course_subject else "",
                        q.question_type,
                        q.question_subtype,
                        q.difficulty,
                        q.test_type,
                    ])
                writer.writerow([])  # blank line
                writer.writerow([])  # second blank line

            return response

        # ===== XLSX EXPORT =====
        elif export_format == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "Duplicates"
            ws.append(["Srno", "Course", "Subject", "Question Type", "Subtype", "Difficulty", "Test Type"])

            for _, questions in grouped.items():
                if len(questions) < 2:
                    continue
                for q in questions:
                    ws.append([
                        q.srno,
                        q.course_subject.course.name if q.course_subject else "",
                        q.course_subject.subject.name if q.course_subject else "",
                        q.question_type,
                        q.question_subtype,
                        q.difficulty,
                        q.test_type,
                    ])
                ws.append([])  # blank line
                ws.append([])  # blank line

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"duplicate_questions_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
            response["Content-Disposition"] = f'attachment; filename={filename}'
            wb.save(response)
            return response

        # ===== JSON fallback =====
        from .serializers import QuestionListSerializer
        serializer = QuestionListSerializer(duplicate_questions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



    @action(
    detail=False,
    methods=['get'],
    permission_classes=[IsAdmin],
    url_path='logs-and-daily-count'
)
    def logs_and_daily_count(self, request):
        """
        Provides daily or date-range question counts, suggestions, and logs.
        Filters: ?date=YYYY-MM-DD OR ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD
                ?srno=123
                ?name=John
        Supports pagination on logs.
        """

        # ✅ Date filters
        date_str = request.query_params.get("date")
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")

        if date_str:
            try:
                filter_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)
            start_date = end_date = filter_date
        else:
            try:
                if start_date_str and end_date_str:
                    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
                    end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
                elif start_date_str:
                    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
                    end_date = start_date
                elif end_date_str:
                    end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
                    start_date = end_date
                else:
                    end_date = timezone.now().date()
                    start_date = end_date - datetime.timedelta(days=6)
            except ValueError:
                return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=400)

        # ✅ srno filter
        srno = request.query_params.get("srno")
        srno_filter = Q()
        if srno:
            srno_filter = Q(srno=srno)

        # ✅ user filter (by name)
        name = request.query_params.get("name")
        name_filter = Q()
        if name:
            name_filter = Q(created_by__name__icontains=name)

        # ✅ Daily "added" counts grouped by date
        daily_added = (
            Question.objects.filter(created_at__date__range=[start_date, end_date])
            .filter(srno_filter & name_filter)
            .values("created_at__date")
            .annotate(total_added=Count("id"))
            .order_by("created_at__date")
        )

        # ✅ Daily "updated" counts grouped by date
        daily_updated = (
            QuestionLog.objects.filter(
                action="EDIT", timestamp__date__range=[start_date, end_date]
            )
            .filter(Q(question__srno=srno) if srno else Q())
            .filter(Q(user__name__icontains=name) if name else Q())
            .values("timestamp__date")
            .annotate(total_updated=Count("id"))
            .order_by("timestamp__date")
        )
        updated_map = {item["timestamp__date"]: item["total_updated"] for item in daily_updated}

        # ✅ Daily "suggestions" counts grouped by date
        # ✅ Daily "suggestions" counts grouped by date
        # ✅ Daily "suggestions" counts grouped by date
        daily_suggestions = (
            Suggestion.objects.filter(created_at__date__range=[start_date, end_date])
            .filter(Q(question__srno=srno) if srno else Q())
            .filter(Q(created_by__name__icontains=name) if name else Q())
            .values("created_at__date")
            .annotate(total_suggestions=Count("id"))
            .order_by("created_at__date")
        )

        # ✅ Build suggestions_map so it's available later
        suggestions_map = {
            item["created_at__date"]: item["total_suggestions"]
            for item in daily_suggestions
        }



        # ✅ Combine day-wise summary
        day_wise_summary = []
        for added in daily_added:
            day = added["created_at__date"]
            day_wise_summary.append({
                "date": day,
                "total_added": added["total_added"],
                "total_updated": updated_map.get(day, 0),
                "total_suggestions": suggestions_map.get(day, 0),
            })

        # Add missing days where only updates happened
        for updated in daily_updated:
            day = updated["timestamp__date"]
            if not any(s["date"] == day for s in day_wise_summary):
                day_wise_summary.append({
                    "date": day,
                    "total_added": 0,
                    "total_updated": updated["total_updated"],
                    "total_suggestions": suggestions_map.get(day, 0),
                })

        # Add missing days where only suggestions happened
        for suggestion in daily_suggestions:
            day = suggestion["created_at__date"]
            if not any(s["date"] == day for s in day_wise_summary):
                day_wise_summary.append({
                    "date": day,
                    "total_added": 0,
                    "total_updated": 0,
                    "total_suggestions": suggestion["total_suggestions"],
                })

        day_wise_summary = sorted(day_wise_summary, key=lambda x: x["date"], reverse=True)

        daily_summary = {
            "date_range": f"{start_date} to {end_date}",
            "total_added": sum(d["total_added"] for d in day_wise_summary),
            "total_updated": sum(d["total_updated"] for d in day_wise_summary),
            "total_suggestions": sum(d["total_suggestions"] for d in day_wise_summary),
            "day_wise": day_wise_summary,
        }

        # ✅ Logs with filters
        logs_qs = (
            QuestionLog.objects.select_related("question", "user")
            .filter(timestamp__date__range=[start_date, end_date])
            .filter(Q(question__srno=srno) if srno else Q())
            .filter(Q(user__name__icontains=name) if name else Q())
            .order_by("-timestamp")
        )

        # ✅ Pagination
        paginator = StandardResultsSetPagination()
        paginated_logs = paginator.paginate_queryset(logs_qs, request)

        logs_data = [
            {
                "id": log.id,
                "question_id": log.question.id,
                "srno": log.question.srno,
                "user": log.user.name,
                "action": log.get_action_display(),
                "timestamp": log.timestamp,
                "ip_address": log.ip_address,
            }
            for log in paginated_logs
        ]

        return paginator.get_paginated_response({
            "daily_question_count": daily_summary,
            "logs": logs_data,
        })



    @permission_classes([IsAdminOrContentDeveloper])
    @action(detail=False, methods=['get'], url_path='question-count')
    def question_count_by_course_subject(self, request):
        """
        Returns question count grouped by Course and Subject,
        with counts of self-practice, full-length,
        and active/inactive questions.
        Groups results by course so same course rows come together.
        """
        data = []

        # Fetch all course_subjects with related course + subject
        course_subjects = CourseSubjects.objects.select_related("course", "subject").order_by("course__name", "subject__name")

        for cs in course_subjects:
            qs = Question.objects.filter(course_subject=cs)

            data.append({
                "course": cs.course.name,
                "subject": cs.subject.name,
                "total_questions": qs.count(),
                "active_questions": qs.filter(is_active=True).count(),
                "inactive_questions": qs.filter(is_active=False).count(),
                "total_self_practice_questions": qs.filter(test_type=Question.SELF_PRACTICE_TEST_TYPE).count(),
                "total_full_length_questions": qs.filter(test_type=Question.FULL_LENGTH_TEST_TYPE).count(),
            })

        return Response(data)




    @action(detail=False, methods=['get'], url_path='download-report')
    def download_report(self, request):
        questions = Question.objects.all().select_related(
            'course_subject__course', 'course_subject__subject', 'topic', 'sub_topic'
        )

        # Prepare workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        grouped_data = {}
        for q in questions:
            course_name = q.course_subject.course.name
            subject_name = q.course_subject.subject.name
            topic_name = getattr(q.topic, 'name', "N/A")
            sub_topic_name = getattr(q.sub_topic, 'name', "N/A")
            difficulty = q.difficulty
            test_type = 'FLT' if q.test_type == Question.FULL_LENGTH_TEST_TYPE else 'Practice'

            key = (course_name, subject_name, topic_name, sub_topic_name, difficulty, test_type)
            grouped_data[key] = grouped_data.get(key, 0) + 1

        # Sort and group by course
        sorted_items = sorted(
            grouped_data.items(),
            key=lambda x: (x[0][0], x[0][1], x[0][2], x[0][3], x[0][4], x[0][5])
        )

        # Separate data by course
        course_data = {}
        for (course, subject, topic, sub_topic, difficulty, test_type), count in sorted_items:
            if course not in course_data:
                course_data[course] = []
            course_data[course].append([subject, topic, sub_topic, difficulty, test_type, count])

        # Create sheet per course
        for course, rows in course_data.items():
            ws = wb.create_sheet(title=course[:31])  # Excel sheet name max length = 31
            ws.append(['Subject', 'Topic', 'Sub Topic', 'Difficulty', 'Test Type', 'Question Count'])
            for row in rows:
                ws.append(row)
            ws.append([])
            ws.append(['Total Questions', '', '', '', '', sum(r[-1] for r in rows)])

        # Prepare response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"questions_report_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response




    @permission_classes([IsAdminOrContentDeveloper])
    def create(self, request, *args, **kwargs):
        data = request.data
        data['created_by'] = request.user.id
        data['updated_by'] = request.user.id
        data['is_active'] = request.user.role.name == 'admin'

        course_subject_id = data.get('course_subject')
        test_type = data.get('test_type')

        max_srno = Question.objects.filter(course_subject_id=course_subject_id, test_type=test_type)\
            .aggregate(Max('srno'))['srno__max'] or 0
        data['srno'] = max_srno + 1

        context = {'request': request}
        serializer = CreateQuestionSerializer(data=data, context=context)

        try:
            serializer.is_valid(raise_exception=True)
            question = serializer.save()

            # ✅ Log the addition
            QuestionLog.objects.create(
                question=question,
                user=request.user,
                action='ADD',
                ip_address=self.get_client_ip(request)

            )

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return get_error_response_for_serializer(logger=self.logger, serializer=serializer, data=request.data)

    @permission_classes([IsAdminOrContentDeveloper])
    @action(detail=False, methods=['post'], url_path='create-multiple')
    def create_multiple_questions(self, request, *args, **kwargs):
        data = request.data
        common_description = data.get('description')
        common_options = data.get('options')
        common_reading_comprehension_passage = data.get('reading_comprehension_passage', None)
        common_question_type = data.get('question_type')
        common_directions = data.get('directions', None)
        common_explanation = data.get('explanation', None)
        questions_data = data.get('questions_data', [])  # List of dictionaries for each course_subject
        question_subtype = data.get('question_subtype')
        
        context = {'request': request, 'question_type': common_question_type, 'question_subtype': question_subtype}

        created_questions = []
        srno_tracker = {}
        try:
            # Fetch global max srno once
            if 'global' not in srno_tracker:
                max_srno = Question.objects.aggregate(Max('srno'))['srno__max'] or 0
                srno_tracker['global'] = max_srno

            for question_data in questions_data:
                srno_tracker['global'] += 1  # increment globally

                individual_data = {
                    'course_subject': question_data.get('course_subject'),
                    'description': common_description,
                    'options': common_options,
                    'reading_comprehension_passage': common_reading_comprehension_passage,
                    'question_type': common_question_type,
                    'question_subtype': question_subtype,
                    'difficulty': question_data.get('difficulty'),
                    'test_type': question_data.get('test_type'),
                    'topic': question_data.get('topic'),
                    'sub_topic': question_data.get('sub_topic', None),
                    'created_by': request.user.id,
                    'updated_by': request.user.id,
                    'is_active': question_data.get('is_active', True if request.user.role.name == 'admin' else False),
                    'show_calculator': question_data.get('show_calculator', False),
                    'directions': common_directions,
                    'explanation': common_explanation,
                    'srno': srno_tracker['global']  # unique across all
                }

                serializer = CreateQuestionSerializer(data=individual_data, context=context)
                serializer.is_valid(raise_exception=True)
                created_question = serializer.save()
                created_questions.append(created_question)

                # ✅ Log each question creation
                QuestionLog.objects.create(
                    question=created_question,
                    user=request.user,
                    action='ADD',
                    ip_address=self.get_client_ip(request)
                )

            return Response(
                CreateQuestionSerializer(created_questions, many=True, context=context).data,
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            self.logger.error(f'Error in create_multiple_questions: {e}')
            return get_error_response_for_serializer(logger=self.logger, serializer=serializer, data=request.data)


    @permission_classes([IsAdmin])
    def update(self, request, pk=None, *args, **kwargs):
        instance = Question.objects.get(id=pk)
        context = {'request': request}
        serializer = self.get_serializer(instance, data=request.data, partial=True, context=context)

        try:
            serializer.is_valid(raise_exception=True)
            updated_question = serializer.save(updated_by=request.user, updated_at=timezone.now())

            # ✅ Log the edit
            QuestionLog.objects.create(
                question=updated_question,
                user=request.user,
                action='EDIT',
                ip_address=self.get_client_ip(request)
            )

            return Response(serializer.data)
        except Exception as e:
            return get_error_response_for_serializer(logger=self.logger, serializer=serializer, data=request.data)


    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user, updated_at=timezone.now())

    @permission_classes([IsAdminOrContentDeveloperOrFaculty])
    def retrieve(self, request, pk=None, *args, **kwargs):
        instance = Question.get_question_by_id(question_id=pk)
        
        serializer = QuestionListSerializer(
            instance=instance,
            context={
                'request': request,
                'test_submission_id': request.GET.get('test_submission_id'),
                'practice_test_result_id': request.GET.get('practice_test_result_id'),
            }
        )

        topics = Topic.objects.filter(course_subject_id=instance.course_subject)
        topics_serializer = TopicSerializer(topics, many=True)
        return Response({'detail': serializer.data, 'topics': topics_serializer.data})


    @permission_classes([IsAdminOrContentDeveloper])
    def destroy(self, request, pk=None, *args, **kwargs):
        instance = Question.get_question_by_id(question_id=pk)
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_destroy(self, instance):
        # instance.is_active = False
        # instance.updated_at = timezone.now()
        # instance.save()
        instance.delete()

    @action(detail=True, methods=['patch'], permission_classes=[IsAdminOrContentDeveloper], url_path='deactivate')
    def deactivate_question(self, request, pk=None):
        question = Question.get_question_by_id(question_id=pk)
        # question.is_active = False
        # question.updated_at = timezone.now()
        # question.save()
        question.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['patch'], permission_classes=[IsAdmin], url_path='activate')
    def activate_question(self, request, pk=None):
        question = Question.get_question_by_id(question_id=pk)
        question.is_active = True
        question.updated_at = timezone.now()
        question.save()
        return Response(status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'], permission_classes=[IsAdminOrContentDeveloper], url_path='soft-deactivate')
    def soft_deactivate_question(self, request, pk=None):
        """
        Soft deactivate a question (set is_active=False instead of deleting).
        """
        try:
            question = Question.get_question_by_id(question_id=pk)
            question.is_active = False
            question.updated_at = timezone.now()
            question.updated_by = request.user
            question.save()
            return Response({"message": "Question deactivated successfully."}, status=status.HTTP_200_OK)
        except Question.DoesNotExist:
            return Response({"error": "Question not found."}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'], permission_classes=[IsAdminOrContentDeveloperOrFacultyOrStudent],
        url_path='details')
    def get_questions_details(self, request):
        question_ids = request.data.get('question_ids', [])

        try:
            question_ids = [int(id) for id in question_ids]
        except ValueError:
            return Response({"error": "Invalid question ID in the list, must be all integers."},
                            status=status.HTTP_400_BAD_REQUEST)

        if not question_ids:
            return Response({"error": "No question IDs provided."}, status=status.HTTP_400_BAD_REQUEST)

        questions = Question.get_questions_for_ids(question_ids)
        question_dict = {question.id: question for question in questions}

        # Non-RC first
        ordered_questions = [
            question_dict[qid] for qid in question_ids
            if qid in question_dict and question_dict[qid].question_subtype != Question.MCQ_READING_COMPREHENSION
        ]

        # RC last
        reading_comprehension_questions = [
            question_dict[qid] for qid in question_ids
            if qid in question_dict and question_dict[qid].question_subtype == Question.MCQ_READING_COMPREHENSION
        ]

        ordered_questions.extend(reading_comprehension_questions)

        serializer = QuestionListSerializer(ordered_questions, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class MaterialViewSet(viewsets.ModelViewSet):
    queryset = Material.objects.all()
    serializer_class = MaterialSerializer
    logger = logging.getLogger('Material')
    aws_storage_client = AwsStorageClient(logger=logger)
    source = 'study_material'

    @permission_classes([IsAdminOrContentDeveloper])
    def update(self, request, pk=None):
        instance = self.get_object()
        data = request.data.copy()
        data['updated_by'] = request.user.id

        # Handle file upload if new file provided
        if data.get('url') is None and 'file' in request.FILES:
            uploaded_file = request.FILES['file']
            fs = FileSystemStorage()
            temp_file_name = uploaded_file.name.split('.')
            file_name = temp_file_name[0] + '_' + str(uuid.uuid4().hex)[:6] + '.' + temp_file_name[1]
            saved_file_name = fs.save(file_name, uploaded_file)
            full_path_to_file = f"{settings.MEDIA_ROOT}/{saved_file_name}"

            # Upload to S3
            self.aws_storage_client.upload_file_from_fs(
                source=self.source,
                filename=saved_file_name,
                full_path_to_file=full_path_to_file,
                content_type=uploaded_file.content_type
            )

            # Delete old file from S3 if exists
            if instance.uploaded_file_name:
                self.aws_storage_client.delete_file(source=self.source, filename=instance.uploaded_file_name)

            # Clean up
            fs.delete(saved_file_name)

            # Update file fields
            data['file_name'] = uploaded_file.name
            data['uploaded_file_name'] = saved_file_name
        elif data.get('url'):
            # If updating with a new URL, remove any previously uploaded file
            if instance.uploaded_file_name:
                self.aws_storage_client.delete_file(source=self.source, filename=instance.uploaded_file_name)
                data['file_name'] = None
                data['uploaded_file_name'] = None

        serializer = MaterialSerializer(instance, data=data, partial=True)

        try:
            serializer.is_valid(raise_exception=True)
            serializer.save()
        except Exception as e:
            if isinstance(e, ValidationError):
                return get_error_response_for_serializer(logger=self.logger, serializer=serializer, data=data)
            else:
                self.logger.exception(f'Error while updating material: {e}')
                return get_error_response(message='Unexpected error occurred while updating material.')

        return Response({'detail': 'Material updated successfully'}, status=status.HTTP_200_OK)    

    @permission_classes([IsAdminOrContentDeveloper])
    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        data['created_by'] = request.user.id
        data['updated_by'] = request.user.id

        if data.get('url', None) is None:
            file_uploaded = request.FILES['file']
            fs = FileSystemStorage()  # defaults to MEDIA_ROOT
            temp_file_name = file_uploaded.name.split('.')
            file_name = temp_file_name[0] + '_' + str(uuid.uuid4().hex)[:6] + '.' + temp_file_name[1]
            saved_file_name = fs.save(file_name, file_uploaded)
            full_path_to_file = settings.MEDIA_ROOT + '/' + saved_file_name

            data['file_name'] = file_uploaded.name
            data['uploaded_file_name'] = saved_file_name
            self.aws_storage_client.upload_file_from_fs(source=self.source, filename=saved_file_name,
                                                        full_path_to_file=full_path_to_file,
                                                        content_type=file_uploaded.content_type)
            fs.delete(saved_file_name)

        serializer = MaterialSerializer(data=data)
        try:
            serializer.is_valid(raise_exception=True)
            serializer.save()
        except Exception as e:
            if isinstance(e, ValidationError):
                return get_error_response_for_serializer(logger=self.logger, serializer=serializer, data=data)
            else:
                self.logger.exception(f'Exception {e}')
                return get_error_response(message='An unexpected error occurred.')

        return Response({'detail': 'File uploaded successfully'}, status=status.HTTP_200_OK)

    @permission_classes([IsAdminOrContentDeveloperOrFacultyOrStudent])
    def list(self, request, *args, **kwargs):
        user = request.user
        course_subject_filter = self.request.query_params.get('course_subject_id', None)

        if user.role.name == 'student':
            course_subject = CourseSubjects.get_course_subject_by_id(course_subject_filter)
            course_enrollment = CourseEnrollment.get_student_enrollment_using_student_course(student_id=user.id,
                                                                                             course_id=course_subject.course_id)
            if course_enrollment.subscription_type == CourseEnrollment.FREE:
                qs = self.queryset.filter(access_type=Material.FREE_ACCESS_TYPE)
            else:
                qs = self.queryset.all()
        else:
            qs = self.queryset.all()

        if course_subject_filter is not None:
            qs = qs.filter(course_subject_id=course_subject_filter)

        # Apply dynamic filtering
        filter_backends = [DjangoFilterBackend]
        filterset = MaterialFilter(request.GET, queryset=qs, request=request)
        if not filterset.is_valid():
            return get_error_response('Invalid filter parameters')

        materials = filterset.qs

        # Apply pagination
        paginator = CustomPageNumberPagination()
        paginator.page_size = 15
        paginated_materials = paginator.paginate_queryset(materials, request)

        serializer = MaterialListSerializer(paginated_materials, many=True)

        # Return the paginated response
        return paginator.get_paginated_response(serializer.data)

    @permission_classes([IsAdminOrContentDeveloperOrFacultyOrStudent])
    def retrieve(self, request, pk=None, *args, **kwargs):
        material = Material.get_material_by_id(material_id=pk)
        serializer = MaterialDetailsSerializer(material)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @permission_classes([IsAdminOrContentDeveloper])
    def destroy(self, request, pk=None, *args, **kwargs):
        instance = Material.get_material_by_id(material_id=pk)
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_destroy(self, instance):
        self.aws_storage_client.delete_file(source=self.source, filename=instance.uploaded_file_name)
        instance.delete()

    @action(detail=True, methods=['patch'], permission_classes=[IsAdminOrContentDeveloper], url_path='deactivate')
    def deactivate_material(self, request, pk=None):
        material = Material.get_material_by_id(material_id=pk)
        self.perform_destroy(material)
        return Response(status=status.HTTP_204_NO_CONTENT)
    